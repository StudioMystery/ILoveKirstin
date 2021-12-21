import os
import re
import time
from tkinter import *
import tkinter.filedialog as fd
import openpyxl
from openpyxl.utils.cell import column_index_from_string
from openpyxl.styles.fills import PatternFill
from openpyxl.styles import colors
import shutil
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.events import EventFiringWebDriver, AbstractEventListener
#from selenium.webdriver.support.events import WebDriverDecorator, WebDriverListener ^replaces above
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

#Notes:
#Needs: Rate-limiting solution + VPN support + live site demo and setup process
#   ^- https://stackoverflow.com/questions/65128879/how-to-bypass-being-rate-limited-html-error-1015-using-python
#   ^- https://pypi.org/project/undetected-chromedriver/#important-note
#
#Remaining Issue: Test IFCUSTOMJS, and write remaining documentation.

#Event Listener Class
class e_listener(AbstractEventListener):
    clicked = False
    def before_click(self, element, driver):
        print ("Event : before element click()")

    def after_click(self, element, driver):
        print ("Event : after element click()")

#Main Webcheck Function
def main_webcheck_function(file_p, base_webpage, header_row, search_column, given_name, window):
    print("webchecking: ", file_p, base_webpage, header_row, search_column, given_name, window)
    main_values_types_list = [file_p, base_webpage, header_row, search_column, given_name]
    for main_val in main_values_types_list:
        pass
        #print(main_val, type(main_val))
#duplicate the file
    parent_file = os.path.abspath(os.path.join(file_p, os.pardir))
    new_filename = str(parent_file) + "\\" + "WEB_" + str(given_name) + ".xlsx"
    shutil.copyfile(file_p, new_filename)
#
#open the duplicate
    wb = openpyxl.load_workbook(new_filename, data_only=True)
    sht_uno = wb[wb.sheetnames[0]]
#
#use the header row to get value of all columns + column number. These become the "page coordinates" that Kirstin adds through web-driver.
    page_coordinates_dict = {}
    for col_ii in range(1, (sht_uno.max_column)):
        add_coord_value = sht_uno.cell(header_row, col_ii).value
        page_coordinates_dict.update({add_coord_value: [col_ii, ]})
    #print(page_coordinates_dict)
#
#With page coordinates list, open a new tkinter window that will ask Kirstin to create a "webcheck scheme" that tells the browser how to get the site values.
    #window1 = Tk()
    window1 = Toplevel(window)
#
#TK Webcheck Select Controls
    wc_search_paragraph = Label(window1, justify=LEFT, text="First, click something in chrome. Then, click the 'Add Search Field' button.")
    wc_search_paragraph.place(x = 4, y = 4)
    #
    wc_current_search_paragraph = Label(window1, justify=LEFT, text="none")
    wc_current_search_paragraph.place(x = 4, y = 25)
    #
    wc_coord_info_paragraph = Label(window1, justify=LEFT, text="Click an element in Chrome. Select a Node. Then, click the 'Add Click Data' button.")
    wc_coord_info_paragraph.place(x = 4, y = 100)
    #
    wc_current_coord_paragraph = Label(window1, justify=LEFT, text="none")
    wc_current_coord_paragraph.place(x = 4, y = 125)
    #
    wc_custom_node_e = Entry(window1)
    wc_custom_node_e.place(x = 125, y = 180)
    #
    #Buttons:
    wc_add_search_field_b = Button(window1, text = "Add Search Field", command = lambda: selectSearchField(e_driver, wc_current_search_paragraph, wc_listbox), bg = "pink", fg = "black")
    #
    wc_view_node_b = Button(window1, text = "View Node", command = lambda: viewNode(e_driver, wc_current_coord_paragraph, wc_listbox, wc_run_status_paragraph), bg = "teal", fg = "white")
    wc_add_css_data_b = Button(window1, text = "Add CSS Data", command = lambda: addCSSData(e_driver, wc_current_coord_paragraph, wc_listbox), bg = "blue", fg = "white")
    wc_clear_css_data_b = Button(window1, text = "Clear CSS Data", command = lambda: clearCSSData(wc_current_coord_paragraph, wc_listbox), bg = "gray", fg = "white")
    wc_create_custom_node_b = Button(window1, text = "Create Custom Node", command = lambda: createCustomNode(wc_current_coord_paragraph, wc_listbox, wc_custom_node_e), bg = "orange", fg = "black")
    wc_delete_node_b = Button(window1, text = "Delete Node", command = lambda: deleteNode(wc_current_coord_paragraph, wc_listbox), bg = "red", fg = "white")
    wc_how_to_b = Button(window1, text = "How To", command = lambda: showHowTo(), bg = "gray", fg = "white")
    #
    webcheck_b_list = [[wc_add_search_field_b, 1, 50], [wc_view_node_b, 1, 150], [wc_add_css_data_b, 75, 150], [wc_clear_css_data_b, 160, 150], [wc_create_custom_node_b, 1, 180], [wc_delete_node_b, 1, 210], [wc_how_to_b, 1, 240]]
    #
    for b_i in webcheck_b_list:
        b_i[0].place(x = b_i[1], y = b_i[2])
#
#TK Webcheck Scheme Controls
    wc_scheme_paragraph = Label(window1, justify=LEFT, text="Node Scheme")
    wc_scheme_paragraph.place(x = 450, y = 4)
    wc_listbox = Listbox(window1)
    wc_listbox.place(x = 450, y = 50)
    wc_run_status_paragraph = Label(window1, justify=LEFT, text=" ")
    #
    wc_href_off_b = Button(window1, text = "HREF Off", command = lambda: haltHREF(e_driver), bg = "orange", fg = "black")
    wc_add_inf_b = Button(window1, text = "Inf", command = lambda: applyInf(e_driver), bg = "gray", fg = "white")
    wc_run_b = Button(window1, text = "Run", command = lambda: runAutoCheck(e_driver, wc_listbox, sht_uno, header_row, search_column, wb, new_filename, wc_run_status_paragraph, parent_file, given_name), bg = "green", fg = "white")
    wc_import_b = Button(window1, text = "Import", command = lambda: getImport(window1, wc_listbox), bg = "gray", fg = "white")
    wc_export_b = Button(window1, text = "Export", command = lambda: saveExport(wc_listbox, given_name, parent_file, wc_current_coord_paragraph), bg = "black", fg = "white")
    #
    scheme_b_list = [[wc_href_off_b, 450, 215], [wc_add_inf_b, 513, 215], [wc_run_b, 540, 215], [wc_import_b, 450, 25], [wc_export_b, 500, 25], [wc_run_status_paragraph, 450, 250]]
    #
    for b_i in scheme_b_list:
        b_i[0].place(x = b_i[1], y = b_i[2])
#
#TK Populating the Scheme List
    open_url_node = "*OPEN --- " + str(base_webpage)
    wc_listbox.insert(1, open_url_node)
    for coord_i in page_coordinates_dict:
        #print(coord_i)
        wc_listbox.insert(END, coord_i) 
#
#Selenium opens and Kirstin can operate the webdriver    
    #
    #Mute Selenium DevTools listening:
    #
    options = webdriver.ChromeOptions()
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    #
    #Get values to search by
    #
    search_values_list = []
    for search_i in (header_row, sht_uno.max_row): #?header_row +1
        add_search_value = sht_uno.cell(search_i, column_index_from_string(search_column)).value
        if add_search_value != None:
            search_values_list.append(add_search_value)
        else:
            print("search value passed", sht_uno.cell(search_i, search_column))
            pass
    #
    #Enable Browser Logging
    #
    d = DesiredCapabilities.CHROME
    d['goog:loggingPrefs'] = { 'browser':'ALL' }
    #
    #Define Selenium Driver to ignore constant log
    #
    sel_service = Service("Chrome 95.048 Webdriver/chromedriver.exe")
    driver = webdriver.Chrome(service=sel_service, options=options, desired_capabilities=d)
    #
    #Add events wrapper for driver
    #
    e_driver = EventFiringWebDriver(driver, e_listener())
    e_driver.get(base_webpage)
    #
    #Remove other windows that open
    #
    if len(e_driver.window_handles) > 1:
        for handle in e_driver.window_handles:
            e_driver.switch_to.window(handle)
            if e_driver.current_url != base_webpage:
                e_driver.close()
            else:
                pass
    #
    #Add event listener that adds the element a user clicks to the chrome Console.
    #
    e_driver.execute_script(open("ILKWebcheckScript.js").read())
    #e_driver.execute_script("document.addEventListener('click', function (event) {var clickedElem = getCSSPath(event.target);console.log(clickedElem)})")
    #
    ###
    #https://stackoverflow.com/questions/48836868/extract-selector-from-event-object-of-current-event-target-getting-object-name
    #http://www.appliedselenium.com/2019/04/event-listener-in-selenium/
    #https://gomakethings.com/checking-event-target-selectors-with-event-bubbling-in-vanilla-javascript/
    #website that Jason's scraper failed on because of request limits: https://www.charlestoncounty.org/tax-info.php
    ###
#
#TK End tags    
    my_icon = PhotoImage(file='alt-ilk-app-icon.png')#icon image
    window1.iconphoto(False, my_icon)
    #
    window1.title('Webcheck')
    window1.geometry("600x300+20+10")
    window1.mainloop()
#

#Additional Functions
def browse_file(window1):
    currdir = os.getcwd()
    tempdir = fd.askopenfilename(parent=window1, initialdir=currdir, title='Please choose your .xlsx file')
    if len(tempdir) > 0:
        return(tempdir)
#
def getConsoleLog(e_driver, target_paragraph):
    runtime_entries = ["Please click something in the browser before attempting to add a selector."]
    #
    for entry in e_driver.get_log('browser'):
        #trim all the excess console log formatting off the css selector in the console log.
        e_m = re.split("\s", entry["message"], 2)
        #add it to a list, in case there are multiple entries
        runtime_entries.append(e_m[2])
    #use the last entry, no matter what
    cur_entry = runtime_entries[-1]
    target_paragraph.configure(text=cur_entry)
    return cur_entry.strip('\"')
#
def getCoordValue(e_driver, css_sel):
    #Given CSS Selector, find value of the element.
    c_sel_stripped = css_sel
    coord_value = e_driver.find_element(By.CSS_SELECTOR,c_sel_stripped).get_attribute('innerText')
    return coord_value
#
def selectSearchField(e_driver, target_paragraph, target_listbox):
    node_index = 1
    node_coord = getConsoleLog(e_driver, target_paragraph)
    node_payload = "*SEARCH --- " + str(node_coord)
    #
    target_lb_item = target_listbox.get(node_index)
    print(target_lb_item)
    if "*SEARCH --- " in target_lb_item:
        print("previous entry detected")
        target_listbox.delete(node_index)
        target_listbox.insert(node_index, node_payload)
    else:
        print("no previous entry detected")
        target_listbox.insert(node_index, node_payload)
    #
    tp_string = str(node_coord)
    target_paragraph.configure(text=tp_string)
#
def viewNode(e_driver, target_paragraph, target_listbox, wc_run_status_paragraph):
    #display the full node in the paragraph
    for item in target_listbox.curselection():
        target_lb_item = target_listbox.get(item)
    css_sele_split_list = re.split(" --- ", target_lb_item, 1)
    css_sele = css_sele_split_list[1]
    if "OPEN" in css_sele_split_list[0]:
        wc_run_status_paragraph.configure(text=css_sele)
    else:
        target_paragraph.configure(text=target_lb_item)
        status_value = getCoordValue(e_driver, css_sele)
        wc_run_status_paragraph.configure(text=status_value)
#
def addCSSData(e_driver, target_paragraph, target_listbox):
    #get the last element clicked + the selected list item
    for item in target_listbox.curselection():
        node_index = str(item)
    #
    print("node index", node_index)
    node_coord = getConsoleLog(e_driver, target_paragraph)
    node_value = getCoordValue(e_driver, node_coord)
    #
    target_lb_item = target_listbox.get(item)
    print(target_lb_item)
    if " --- " in target_lb_item:
        print("previous entry detected")
        node_name_split_list = re.split(" --- ", target_listbox.get(item), 2)
        node_name = node_name_split_list[0]
    else:
        print("no previous entry detected")
        node_name = target_lb_item
    #
    node_payload = str(node_name) + " --- " + str(node_coord)
    #
    target_listbox.delete(node_index)
    target_listbox.insert(node_index, node_payload)
    #
    tp_string = str(node_value) + " - " + str(node_coord)
    target_paragraph.configure(text=tp_string)
#
def clearCSSData(target_paragraph, target_listbox):
    #remove css selector from node
    #get the last element clicked + the selected list item
    for item in target_listbox.curselection():
        node_index = str(item)
    #
    print("node index", node_index)
    #node_coord = getConsoleLog(e_driver, target_paragraph)
    #
    target_lb_item = target_listbox.get(item)
    print(target_lb_item)
    if " --- " in target_lb_item:
        print("previous entry detected")
        node_name_split_list = re.split(" --- ", target_listbox.get(item), 2)
        node_name = node_name_split_list[0]
    else:
        print("no previous entry detected")
        node_name = target_lb_item
    #
    node_payload = str(node_name)
    #
    target_listbox.delete(node_index)
    target_listbox.insert(node_index, node_payload)
    #
    target_paragraph.configure(text="CSS Selector Removed.")
#
def createCustomNode(target_paragraph, target_listbox, wc_custom_node_e):
    #add a new node to the node scheme
    for item in target_listbox.curselection():
        node_index = str(item)
    node_payload = wc_custom_node_e.get()
    target_listbox.insert(node_index, node_payload)
    target_paragraph.configure(text="Node Created")
#
def deleteNode(target_paragraph, target_listbox):
    #remove an existing node from the node scheme
    for item in target_listbox.curselection():
        node_index = str(item)
    target_listbox.delete(node_index)
    target_paragraph.configure(text="Node Deleted")
#
def showHowTo():
    #adds how to guide information for using webcheck
    window2 = Toplevel(height=600, width=750)
    i_t_1 = "Hi Kirstin! Love you babe!"
    i_t_2 = "I hope you are enjoying your application!"
    i_t_3 = "This page explains how to use the 'Create Custom Node' button to do all kinds of things. Regular nodes are imported from the excel file you choose. Each node is named after a column in the excel file, and is supposed to be tied directly to an HTML element(via css selector)."
    i_t_4 = "However, you don't have to limit yourself to these nodes, you can create your own, too!"
    i_t_5 = "Custom nodes can do lots of different things, but they follow a certain syntax: * + Name + [ + Variable + ] + --- + CSS Selector/URL."
    i_t_6 = '\"*OPEN --- \"   this node opens a url, placed after the \" --- \".'
    i_t_7 = '\"*SEARCH --- \"   this node lets you select a field to fill in the search value for every row on your excel sheet.'
    i_t_8 = '\"*CLICK --- \"   this node lets you simulate a click on any element in the UI.'
    i_t_9 = '\"*TEXT[] --- \"   this node lets you send a text string to the selected element.'
    i_t_10 = '\"*COLUMN[] --- \"   this node lets you send a text string from another column in your excel sheet.'
    i_t_11 = '\"*KEYSTROKE[] --- \"   this node allows you to simulate pressing a keyboard key.'
    i_t_12 = '\"*TIMER[] --- \"   this node allows you to pause the program and force it to wait for a given number of seconds. The CSS Selector isn\'t required.'
    i_t_13 = '\"*IFCUSTOMJS[] --- \"    this node allows you to enter custom javascript during the automated webchecking process, IF the CSS Selector is verified(i.e. it exists in the DOM).'
    i_t_14 = "As you can see, there are lots of different acceptable custom nodes. But be careful! Sending an invalid node may not work the way you want! The single-quoted text above lays out the general format of each node, but you still need to select an HTML element to target. In addition, you can type your text / column letter in the '[]' brackets."
    i_t_15 = "Lastly, There are the \"HREF OFF\" and \"Inf\" buttons. HREF OFF essentially turns off any links in the page, preventing the user from unintentionally navigating when attempting to gain selector data by clicking. Users will need to reload the page in order to use links again. Inf stands for Console.info(), the function that replaces Console.log() in the event that a website disables logging in production. If your clicks aren't sending selector data to the nodes in the scheme, try clicking the Inf button, then retry your attempt."
    instructional_text = i_t_1 + "\n" + "\n" + i_t_2 + "\n" + "\n" + i_t_3 + "\n" + "\n" + i_t_4 + "\n" + "\n" + i_t_5 + "\n" + "\n" + i_t_6 + "\n" + i_t_7 + "\n" + i_t_8 + "\n" + i_t_9 + "\n" + i_t_10 + "\n" + i_t_11 + "\n" + i_t_12 + "\n" + i_t_13 + "\n" + "\n" + i_t_14 + "\n" + "\n" + i_t_15
    wc_how_to_paragraph = Label(window2, wraplength= "750", justify=LEFT, text=instructional_text)
    wc_how_to_paragraph.place(x = 4, y = 4)
    wc_how_to_b = Button(window2, text = "Hide How To", command = lambda: window2.destroy(), bg = "gray", fg = "white")
    wc_how_to_b.place(x = 670, y = 10)
#
def browse_file(window1):
    currdir = os.getcwd()
    tempdir = fd.askopenfilename(parent=window1, initialdir=currdir, title='Please choose your SCHEME.txt file')
    if len(tempdir) > 0:
        return(tempdir)
#
def getImport(window1, target_listbox_im):
    #print(type(target_listbox_im.size()))
    #print(target_listbox_im.size())
    for ns_i in range(0, target_listbox_im.size()):
        target_listbox_im.delete(END)
    #
    tempdir = browse_file(window1)
    import_read = open(tempdir, "r")
    import_file = import_read.read()
    import_lines = import_file.count('\n')
    #
    i_m = re.split('\n', import_file, import_lines)
    for t_i in range(0, import_lines):
        target_listbox_im.insert(END, i_m[t_i])
    #
    print("import completed")
    import_read.close()
#
def saveExport(target_listbox, given_name, parent_file, target_paragraph):
    new_ex_filename = str(parent_file) + "\\" + "SCHEME_" + str(given_name) + ".txt"
    scheme_file = open(new_ex_filename, "w")
    for t_i in range(0, target_listbox.size()):
        scheme_file.write(target_listbox.get(t_i) + '\n')
    scheme_file.close()
    target_paragraph.configure(text="Export Completed.")
    print("export complete")
#
def haltHREF(e_driver):
    e_driver.execute_script("var all = document.getElementsByTagName('*');for (var i=0, max=all.length; i < max; i++){all[i].setAttribute('onclick', 'return false');}")
#
def applyInf(e_driver):
    print("inf applied")
    #work_d_code = 'console.dir("tea and crumpits")'
    #e_driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {'source': 'alert("Hooray! I did it!")'})
    e_driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {'source': open("ILKWebcheckScript.js").read()})
    e_driver.refresh()
    #e_driver.execute_script(open("ILKWebcheckScript.js").read())
    e_driver.execute_script("document.addEventListener('click', function (event) {var clickedElem = UTILS.cssPath(event.target);console.info(clickedElem)})")
#
def runAutoCheck(e_driver, target_listbox, sht_uno, header_row, search_column, wb, new_filename, wc_run_status_paragraph, parent_file, given_name):
    #secondary function to run the automated checking process.
    #
    #Open the base url
    error_log = ['']
    excel_row_errors = ['']
    wc_run_status_paragraph.configure(text="Webcheck In-Progress...")
    print("begining automated loop...")
    #
    #Verifty that the scheme has a target url
    if "*OPEN --- " in target_listbox.get(0):
        print("URL Found")
    else:
        print("No URL")
        return False
    #
    #print(sht_uno.max_row)
    #print('')
    #
    #For every row in the excel sheet, run every step in the Node Scheme list.
    for row_iii in range(header_row + 1, (sht_uno.max_row) + 1):
        print('\n' + "Excel Row: " + str(row_iii))
        #for every item in the scheme, eval the item
        for t_i in range(1, target_listbox.size()):
            #in case there are errors, save the errors in a list.
            try:
                print("Node Scheme Item: ", t_i, target_listbox.get(t_i))
                t_i_value = target_listbox.get(t_i)
                # * indicates it's a selenium task
                if "*OPEN --- " in target_listbox.get(t_i):
                    star_open_val = re.split(" --- ", target_listbox.get(0), 1)
                    e_driver.get(star_open_val[1])
                elif "*SEARCH --- " in t_i_value:
                    star_search_split = re.split(" --- ", t_i_value, 1)
                    star_search_coord = star_search_split[1]
                    excl_search_value = sht_uno.cell(row=row_iii, column=column_index_from_string(search_column)).value
                    #
                    editor = WebDriverWait(e_driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, star_search_coord)))
                    #editor.click()
                    editor.send_keys(Keys.CONTROL, 'a')
                    editor.send_keys(Keys.BACKSPACE)
                    editor.send_keys(excl_search_value)
                    editor.send_keys(Keys.RETURN)
                    #
                elif "*CLICK --- " in target_listbox.get(t_i):
                    star_click_split = re.split(" --- ", t_i_value, 1)
                    star_click_coord = star_click_split[1]
                    #
                    editor = WebDriverWait(e_driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, star_click_coord)))
                    editor.click()
                    #
                elif "*TEXT[" in target_listbox.get(t_i):
                    star_text_split = re.split(" --- ", t_i_value, 1)
                    star_text_coord = star_text_split[1]
                    star_text_to_fix = re.search(r"\[([A-Za-z0-9_]+)\]", star_text_split[0])
                    star_text_to_group = star_text_to_fix.group()
                    star_text_to_replace_1 = star_text_to_group.replace("[", "")
                    star_text_to_replace_2 = star_text_to_replace_1.replace("]", "")
                    star_text_value = star_text_to_replace_2
                    print(star_text_value)
                    #
                    editor = WebDriverWait(e_driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, star_text_coord)))
                    editor.send_keys(Keys.CONTROL, 'a')
                    editor.send_keys(Keys.BACKSPACE)
                    editor.send_keys(star_text_value)
                    #
                elif "*COLUMN[" in target_listbox.get(t_i):
                    star_col_split = re.split(" --- ", t_i_value, 1)
                    star_col_coord = star_col_split[1]
                    print(star_col_split[0])                
                    star_col_to_fix = re.search(r"\[([A-Za-z0-9_]+)\]", star_col_split[0])
                    star_col_to_group = star_col_to_fix.group()
                    star_col_to_replace_1 = star_col_to_group.replace("[", "")
                    star_col_to_replace_2 = star_col_to_replace_1.replace("]", "")
                    star_col_value = star_col_to_replace_2
                    #print(sht_uno.cell(row=header_row, column= column_index_from_string(star_col_value)).value)
                    #
                    excl_search_value = sht_uno.cell(row=row_iii, column=column_index_from_string(star_col_value)).value
                    #
                    editor = WebDriverWait(e_driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, star_col_coord)))
                    editor.send_keys(Keys.CONTROL, 'a')
                    editor.send_keys(Keys.BACKSPACE)
                    editor.send_keys(excl_search_value)
                elif "*KEYSTROKE[" in target_listbox.get(t_i):
                    star_key_split = re.split(" --- ", t_i_value, 1)
                    star_key_coord = star_key_split[1]
                    star_key_to_fix = re.search(r"\[([A-Za-z0-9_]+)\]", star_key_split[0])
                    star_key_to_group = star_key_to_fix.group()
                    star_key_to_replace_1 = star_key_to_group.replace("[", "")
                    star_key_to_replace_2 = star_key_to_replace_1.replace("]", "")
                    star_key_value = star_key_to_replace_2
                    #
                    e_driver.implicitly_wait(10)
                    WebDriverWait(e_driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, star_key_coord)))
                    editor = e_driver.find_element_by_css_selector(star_key_coord)
                    editor.send_keys(Keys.__getattribute__(Keys, star_key_value))
                    #
                elif "*TIMER[" in target_listbox.get(t_i):
                    star_timer_split = re.split(" --- ", t_i_value, 1)
                    star_timer_coord = star_timer_split[1]
                    star_timer_to_fix = re.search(r"\[([A-Za-z0-9_]+)\]", star_timer_split[0])
                    star_timer_to_group = star_timer_to_fix.group()
                    star_timer_to_replace_1 = star_timer_to_group.replace("[", "")
                    star_timer_to_replace_2 = star_timer_to_replace_1.replace("]", "")
                    star_timer_value = int(star_timer_to_replace_2)
                    #
                    print("Start Time: %s" % time.ctime())
                    e_driver.implicitly_wait(star_timer_value)
                    time.sleep(star_timer_value)
                    print("End Time: %s" % time.ctime())
                    #WebDriverWait(e_driver, star_timer_value)         
                elif "*IFCUSTOMJS[" in target_listbox.get(t_i):
                    #*IFCUSTOMJS[alert("Hello! I am an alert box!")] --- ul#template-header-nav > li:nth-child(1) > a
                    star_ifcustomjs_split = re.split(" --- ", t_i_value, 1)              
                    star_ifcustomjs_coord = star_ifcustomjs_split[1]                    
                    star_ifcustomjs_to_fix = re.search("IFCUSTOMJS(.*) --- ", t_i_value).group().replace("IFCUSTOMJS[", "").replace("] --- ", "")
                    star_ifcustomjs_value = str(star_ifcustomjs_to_fix)
                    #
                    if_Elements = e_driver.find_elements(By.CSS_SELECTOR, star_ifcustomjs_coord)
                    if len(if_Elements) > 0:
                        e_driver.execute_script(star_ifcustomjs_value)
                    else:
                        print("No elements found matching IFCUSTOMJS selector.")
                #assume the node is referring to a column on the sheet.
                else:              
                    #Error!!! If "--- " doesn't exist, skip the value
                    if " --- " not in target_listbox.get(t_i):
                        pass
                    #The node is ready to check
                    else:
                        #define the vars
                        plain_node_split = re.split(" --- ", t_i_value, 1)
                        plain_node_coord = plain_node_split[1]
                        plain_node_value = plain_node_split[0]
                        excl_node_value = "unk node val"
                        page_elem_value = "unk elem val"
                        #find the matching column for the node
                        for i in range(1,sht_uno.max_column):
                            #if the column name matches the scheme name, then evaluate for the value
                            if sht_uno.cell(row=header_row, column=i).value == plain_node_value:
                                excl_node_value = sht_uno.cell(row=row_iii,column=i).value
                                page_elem_value = getCoordValue(e_driver, plain_node_coord)
                                print("Excel value found: ", excl_node_value)
                                print("Node value found: ", page_elem_value)
                                #if the values match, do nothing, otherwise add elem to duped excel sheet, then highlight the cell in red
                                if page_elem_value == excl_node_value:
                                    print("Match Found, no changes made.")
                                    pass
                                #
                                elif row_iii in excel_row_errors:
                                    print("row error found. ", "Excel row: ", row_iii, " Node Item: ", t_i_value)
                                    orangeFill = PatternFill(patternType='solid', fgColor=colors.Color(rgb='00FF4500'))
                                    sht_uno.cell(row=row_iii,column=i).value = "ERROR!"
                                    sht_uno.cell(row=row_iii,column=i).fill = orangeFill
                                    del(page_elem_value)
                                else:
                                    print("elem/value mismatch found. ", "Excel row: ", row_iii, " Node Item: ", t_i_value)
                                    redFill = PatternFill(patternType='solid', fgColor=colors.Color(rgb='00FF0000'))
                                    sht_uno.cell(row=row_iii,column=i).value = page_elem_value
                                    sht_uno.cell(row=row_iii,column=i).fill = redFill
                                    del(page_elem_value)
                            else:
                                pass
            #If an error occured during the autocheck, add where it happened to a .txt file.             
            except Exception as what_went_wrong:
                excel_row_errors.append(row_iii)
                error_statement = str("ERROR: runAutoCheck encountered an error on Excel Row: " + str(row_iii) + " " + str(what_went_wrong))
                error_log.append(error_statement)
                print(error_statement)
        #Save the file after every cell is validated
        wb.save(new_filename)
    #
    #Check for errors in error log, then create .txt with errors found.
    if error_log != ['']:
        print(error_log)
        new_er_filename = str(parent_file) + "\\" + "ERRORS_" + str(given_name) + " TIME - " + str(time.strftime("%Y%m%d-%H%M%S")) + ".txt"
        er_file = open(new_er_filename, "w")
        for er_i in error_log:
            er_str = er_i + '\n'
            er_file.write(er_str)
        er_file.close()
    else:
        print("No errors logged.")
    #
    #Save the duplicate file, and print the number of highlights.
    wc_run_status_paragraph.configure(text="Webcheck complete.")
    wb.save(new_filename)
    print('\n' + "webcheck automated loop complete.")
#

#This tests the module by itself with basic values from download file
#main_webcheck_function("C:\\Users\\Thadd xSx\\Downloads\\TSR Raw list 8-6-2021 - GRD-MT-Yellowstone_County-Lien-2021-08-31GRD-AVM-BT-FLOOD.xlsx", "https://en.wikipedia.org/wiki/Machine_learning", 10, "F", given_name="testliner")
#file_p, base_webpage, header_row, search_column, given_name="default"