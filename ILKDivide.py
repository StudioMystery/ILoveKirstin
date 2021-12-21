from tkinter import *
import openpyxl
from openpyxl.utils.cell import column_index_from_string
import os

#This is the main function of the module, called by ILK app.
def main_divide_function(file_p, bid_number, given_name, header_row, sort_column):
#
#Start the divide function, set the row/column definitions and format the types.
    print("dividing: ", file_p, bid_number, given_name, header_row, sort_column)
    wb = openpyxl.load_workbook(file_p)
    sht_uno = wb[wb.sheetnames[0]]
    #sht_a21 = sht_uno['A21'].value
    #print("TEST, ", sht_a21)
    #
    master_sort_column = column_index_from_string(sort_column)   
    master_header_row = int(header_row) + 1
    add_master_rows = int(header_row) + 1 #second variable needed because master_header_row changes at some point.
#
#Create a list of sorted numerical values, ordered smallest to largest, from the header row to the last value
    amount_ordered_list = []
    for cell_i in range(master_header_row, (sht_uno.max_row + 1)):
        add_to_sort_list = sht_uno.cell(row=cell_i, column=master_sort_column).value
        amount_ordered_list.append([add_to_sort_list, cell_i])
        cell_i = cell_i + 1
    amount_ordered_list.sort()
    ###use the secondary values of each item in amount_ordered_list to order the row_dict later on.
    desired_order_list = []#pulls the amount exclusively out of the above list 
    for obj in amount_ordered_list:
        desired_order_list.append(obj[1])
    #    print(obj[0], ", ", obj[1])
    #    sumdoc = sumdoc + obj[0]
    #print("len of list", len(amount_ordered_list),", sum: ", sumdoc)
    #input()
    #print(desired_order_list[0])
#
#Grab all rows from the selected sheet and add them to row_dict, keyed off of line number / myint
    row_dict = {}
    mylist = []
    myint = 1
    for row in sht_uno.rows:
        for cell in row:
            mylist.append(cell.value)
        row_dict.update({str(myint):mylist})
        myint = myint + 1
        del mylist
        mylist = []
#
#Create new dict ordered by reorder because the row_dict contains header rows
    #print(row_dict["10"]) #row containing headers. 0 is not a row
    reordered_dict = {}
    for k in desired_order_list:
        #is k the row or the number?
        reordered_dict.update({str(k):row_dict[str(k)]})
    #print(reordered_dict["794"]) #provides row 794, corresponding to item 784, with amount of 2.68
    reo_dict_key_len = total_keys(reordered_dict)
    #print(reo_dict_key_len)
    #input()
    #row_dict = 
#
#Print total rows and add cells in master rows to a list for later
    #dict_key_len = total_keys(row_dict)
    #print("Total Rows = ", dict_key_len - master_header_row + 1)
    add_master_rows_list = []
    for hrow in range(1, int(add_master_rows)):
        add_master_rows_list.append(row_dict[str(hrow)])
#
#Use the new dict to create a list of cells to add to new files
    ib = 1 #bidder iterate
    rb = 0 #master_header_row #row iterate
    cb = 0 #cycle iterate
    printlist = []
    #print("reo_dict_key_len: ", reo_dict_key_len, ", vs desired_list_len: ", len(desired_order_list), ", vs reordered_dict_len: ", len(reordered_dict))
    #prinput("stop")
    for ib in range(int(bid_number)): #for every bid sheet
        while rb < reo_dict_key_len: #for every row
            #print("Iteration: ", rb, ", Correlates To: ", desired_order_list[rb], ", With Cell Value: ", reordered_dict[str(desired_order_list[rb])][:6])
            printlist.append(reordered_dict[str(desired_order_list[rb])])
            rb = rb + int(bid_number)
        #print("Rows to Add = ", len(printlist))
        save_bid_xlsx_sheet(printlist, ib, given_name, file_p, add_master_rows_list) #this is doing the work of adding each list of rows (1, 4, 7, 10) to each bidder's new document.
        del printlist
        printlist = []
        cb = cb + 1
        ib = ib + 1
        rb = cb
        #print("ib: ", ib, ", rb: ", rb, ", cb: ", cb)
        #prinput("stop")
#
#This is used in divide workbook to find number of rows
def total_keys(test_dict):
    return (0 if not isinstance(test_dict, dict) 
    else len(test_dict) + sum(total_keys(val) for val in test_dict.values()))
#
#This manually creates each xlsx file in the dir with the right names
def save_bid_xlsx_sheet(printlist, ib, given_name, file_p, add_master_rows_list):
    parent_file = os.path.abspath(os.path.join(file_p, os.pardir))
    new_filename = str(parent_file) + "\\" + "DIV_" + str(given_name) + str(ib) + ".xlsx"
    wb = openpyxl.Workbook()
    wb.save(new_filename)
    workbook_2 =  openpyxl.load_workbook(str(new_filename))
    sheet = workbook_2.active
    for row in add_master_rows_list:
        sheet.append(row)
    for row in printlist:
        sheet.append(row)
    workbook_2.save(filename=new_filename)
    #print("file created")
    file_p = file_p
#
#This returns a dict that was stripped of the list of keys
def without_keys(dict, list_of_keys):
    return {x: dict[x] for x in dict if x not in list_of_keys}
#
#This is a debugging tool that prints and pauses the program for input.
def prinput(str):
    print(str)
    input()
#
#This tests the module by itself with basic values from download file
#main_divide_function(file_p="C:\\Users\\Thadd xSx\\Downloads\\East Baton Rouge bids list (11-7-2021) - Copy.xlsx", bid_number="3", header_row="10", given_name="testopoly", sort_column="B") #AN is Amount, C:\\Users\\Thadd xSx\\Downloads\\TSR Raw list 8-6-2021 - GRD-MT-Yellowstone_County-Lien-2021-08-31GRD-AVM-BT-FLOOD.xlsx
#