import os
from tkinter import *
import openpyxl
from openpyxl.utils.cell import column_index_from_string
from openpyxl.styles.fills import PatternFill
from openpyxl.styles import colors
import shutil

from openpyxl.xml.constants import MAX_COLUMN

whiteFill = PatternFill(patternType='solid', fgColor=colors.Color(rgb='FFFFFFFF'))

def main_format_function(file_p, hide_list, grow_list, clear_list, clear_all, given_name="default"):
    print("formatting: ", file_p, hide_list, grow_list, clear_list, clear_all, given_name)
#duplicate the file
    parent_file = os.path.abspath(os.path.join(file_p, os.pardir))
    new_filename = str(parent_file) + "\\" + "FMT_" + str(given_name) + ".xlsx"
    shutil.copyfile(file_p, new_filename)
#
#open the duplicate
    wb = openpyxl.load_workbook(new_filename)
    sht_uno = wb[wb.sheetnames[0]]
#
#listify strings
    hide_list = list(hide_list.split(" "))
    grow_list = list(grow_list.split(" "))
    clear_list = list(clear_list.split(" "))
    print(hide_list)
    print(grow_list)
    print(clear_list)
#
#hide a column:
    if hide_list != ['']:
        for i_col in hide_list:
            sht_uno.column_dimensions[i_col].hidden= True
    else:
        pass
#    
#widen a column:
    if grow_list != ['']:
        for ii_col in grow_list:
            sht_uno.column_dimensions[ii_col].width= 50
    else:
        pass
#
#if clear all:
    if clear_all == "YES":
        clear_list = []
        for iii_column in range(1, sht_uno.max_column):
            clear_list.append(iii_column)
    else:
        pass
#
#clear formatting on a column:
    if clear_list != ['']:
        for iiii_col in clear_list:
            sht_uno.column_dimensions[iiii_col].hidden= False
            sht_uno.column_dimensions[iiii_col].width= 13
            for iiii_cell in range(1, sht_uno.max_row):
                sht_uno.cell(row=iiii_cell,column=iiii_col).fill = whiteFill
    else:
        pass
#
#call saveas function to finish the file.
    wb.save(new_filename)
    print("formatting complete")
#
#This tests the module by itself with basic values from download file
#main_format_function(file_p="C:\\Users\\Thadd xSx\\Downloads\\TSR Raw list 8-6-2021 - GRD-MT-Yellowstone_County-Lien-2021-08-31GRD-AVM-BT-FLOOD.xlsx", given_name="testrozini")
#